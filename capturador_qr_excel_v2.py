# -*- coding: utf-8 -*-
'''
Capturador QR → Excel (Tabulado por columnas) — v2
- Cada QR = 1 fila
- Todas las claves en columnas fijas (compatibles con tu CSV/plantilla)
- Soporta 3 modos de tus QRs: Mini-página (data: HTML), Texto legible, JSON (v1 y v3.3)
- Salida en EXCEL (.xlsx) y opcional CSV
- Lector en modo HID (teclado), con ENTER como sufijo del escaneo
'''
import re, json, base64, urllib.parse, html, datetime, sys, os, csv
from pathlib import Path
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from openpyxl import Workbook, load_workbook

COLUMNS = [
    "timestamp",
    "nombre_generico","concentracion","forma_farmaceutica","presentacion",
    "control","registro_sanitario",
    "lote","fecha_fabricacion","fecha_vencimiento",
    "fabricante","importador",
    "id_hospital","fecha_reempaque",
    "conservacion","advertencias","normativa",
    "url_qr","raw"
]

DEFAULT_XLSX = Path("scans_qr_tabulado.xlsx")
DEFAULT_CSV  = Path("scans_qr_tabulado.csv")

def ensure_workbook(path: Path):
    if path.exists():
        try:
            wb = load_workbook(path)
            ws = wb.active
            headers = [c.value for c in ws[1]]
            if headers != COLUMNS:
                ws.delete_rows(1, ws.max_row)
                ws.append(COLUMNS)
            return wb, ws
        except Exception:
            pass
    wb = Workbook()
    ws = wb.active
    ws.title = "Escaneos"
    ws.append(COLUMNS)
    return wb, ws

def normalize_key(s: str) -> str:
    s = (s or "").strip().lower()
    rep = { "á":"a","é":"e","í":"i","ó":"o","ú":"u","ñ":"n","/":"_","-":"_","·":"_","—":"_","–":"_","  ":" " }
    for a,b in rep.items():
        s = s.replace(a,b)
    import re as _re
    s = _re.sub(r"\s+", "_", s)
    s = _re.sub(r"[^a-z0-9_]", "", s)
    return s

KEY_ALIASES = {
    "medicamento": ("nombre_generico","concentracion"),
    "forma_presentacion": ("forma_farmaceutica","presentacion"),
    "formapresentacion": ("forma_farmaceutica","presentacion"),
    "forma": "forma_farmaceutica",
    "presentacion": "presentacion",
    "tipo_de_control": "control",
    "control": "control",
    "registro_sanitario": "registro_sanitario",
    "regsan": "registro_sanitario",
    "regsani": "registro_sanitario",
    "lote": "lote",
    "fecha_fabricacion": "fecha_fabricacion",
    "fabricacion": "fecha_fabricacion",
    "fab": "fecha_fabricacion",
    "fecha_vencimiento": "fecha_vencimiento",
    "vencimiento": "fecha_vencimiento",
    "vence": "fecha_vencimiento",
    "fabricante": "fabricante",
    "importador": "importador",
    "entidad": "id_hospital",
    "id_hospital": "id_hospital",
    "reempaque": "fecha_reempaque",
    "fecha_reempaque": "fecha_reempaque",
    "conservacion": "conservacion",
    "advertencias": "advertencias",
    "normativa": "normativa",
    "url": "url_qr",
    "url_qr": "url_qr",
}

def split_name_conc(text: str):
    text = (text or "").strip()
    if not text: return "",""
    import re as _re
    m = _re.match(r"^(.*?)[\s,;-]+(\d.*)$", text)
    if m: return m.group(1).strip(), m.group(2).strip()
    parts = text.split()
    if len(parts) >= 2: return " ".join(parts[:-1]), parts[-1]
    return text, ""

def split_forma_presentacion(text: str):
    text = (text or "").strip()
    if "·" in text: a,b = text.split("·",1)
    elif "-" in text: a,b = text.split("-",1)
    elif "|" in text: a,b = text.split("|",1)
    else: return text, ""
    return a.strip(" .|-"), b.strip(" .|-")

def parse_text_legible(text: str) -> dict:
    out = {}
    for rawline in text.splitlines():
        line = rawline.strip()
        if not line or ":" not in line: continue
        k, v = line.split(":", 1)
        nk = normalize_key(k); v = v.strip()
        if nk in ("medicamento",):
            ng, cc = split_name_conc(v); out["nombre_generico"], out["concentracion"] = ng, cc
        elif nk in ("forma_presentacion","formapresentacion"):
            f, p = split_forma_presentacion(v); out["forma_farmaceutica"], out["presentacion"] = f, p
        else:
            target = KEY_ALIASES.get(nk)
            if isinstance(target, str): out[target] = v
    if "forma_farmaceutica" in out and "presentacion" not in out:
        f,p = split_forma_presentacion(out["forma_farmaceutica"]); out["forma_farmaceutica"], out["presentacion"] = f,p
    return out

def parse_json_payload(text: str) -> dict:
    data = json.loads(text); out = {}
    if isinstance(data.get("medicamento"), dict):
        m = data.get("medicamento", {})
        out["nombre_generico"] = m.get("nombre_generico",""); out["concentracion"] = m.get("concentracion","")
        out["forma_farmaceutica"] = m.get("forma_farmaceutica",""); out["presentacion"] = m.get("presentacion","")
        out["control"] = m.get("tipo_control","") or data.get("medicamento",{}).get("control","")
        out["registro_sanitario"] = m.get("registro_sanitario","")
        l = data.get("lote", {}); out["lote"] = l.get("codigo",""); out["fecha_fabricacion"] = l.get("fecha_fabricacion",""); out["fecha_vencimiento"] = l.get("fecha_vencimiento","")
        fab = data.get("fabricacion", {}); out["fabricante"] = fab.get("fabricante","") or data.get("fabricante",""); out["importador"] = fab.get("importador","") or data.get("importador","")
        rep = data.get("reempaque", {}); out["id_hospital"] = rep.get("entidad","") or data.get("entidad",{}).get("id",""); out["fecha_reempaque"] = rep.get("fecha","") or data.get("entidad",{}).get("reempaque","")
        norm = data.get("normativa", []); out["normativa"] = "; ".join(norm) if isinstance(norm, list) else str(norm)
    elif isinstance(data.get("producto"), dict):
        p = data.get("producto", {})
        out["nombre_generico"] = p.get("nombre_generico",""); out["concentracion"] = p.get("concentracion","")
        out["forma_farmaceutica"] = p.get("forma_farmaceutica",""); out["presentacion"] = p.get("presentacion","")
        out["registro_sanitario"] = p.get("registro_sanitario",""); out["control"] = p.get("control","")
        l = data.get("lote", {}); out["lote"] = l.get("codigo",""); out["fecha_fabricacion"] = l.get("fecha_fabricacion",""); out["fecha_vencimiento"] = l.get("fecha_vencimiento","")
        out["fabricante"] = data.get("fabricante",""); out["importador"] = data.get("importador","")
        ent = data.get("entidad", {}); out["id_hospital"] = ent.get("id",""); out["fecha_reempaque"] = ent.get("reempaque","")
        norm = data.get("normativa", []); out["normativa"] = "; ".join(norm) if isinstance(norm, list) else str(norm)
    return out

def parse_data_url(text: str) -> dict:
    if text.startswith("data:text/html;base64,"):
        html_txt = base64.b64decode(text.split(",",1)[1]).decode("utf-8","ignore")
    elif text.startswith("data:text/html,"):
        html_txt = urllib.parse.unquote(text.split(",",1)[1])
    else:
        return {}
    import re as _re
    rows = _re.findall(r"<tr>\s*<th[^>]*>(.*?)</th>\s*<td[^>]*>(.*?)</td>\s*</tr>", html_txt, flags=_re.I|_re.S)
    out = {}
    for th, td in rows:
        key = html.unescape(_re.sub("<.*?>","", th)).strip()
        val = html.unescape(_re.sub("<.*?>","", td)).strip()
        nk = normalize_key(key)
        if nk in ("medicamento",):
            ng, cc = split_name_conc(val); out["nombre_generico"], out["concentracion"] = ng, cc
        elif nk in ("forma_presentacion","formapresentacion","forma"):
            f, p = split_forma_presentacion(val); out["forma_farmaceutica"], out["presentacion"] = f, p
        else:
            target = KEY_ALIASES.get(nk)
            if isinstance(target, str): out[target] = val
    out.setdefault("normativa", "Res 1478/2006 · Circ 01/2016 FNE")
    return out

def parse_scan(text: str) -> dict:
    text = (text or "").strip()
    if not text: return {}
    try:
        if text.startswith("data:text/html"):
            data = parse_data_url(text)
        elif text.startswith("{"):
            data = parse_json_payload(text)
        else:
            data = parse_text_legible(text)
    except Exception:
        data = {}
    return data

class App:
    def __init__(self, root):
        self.root = root
        root.title("QR → Excel (Tabulado por columnas) — v2")
        root.geometry("860x520")
        root.attributes("-topmost", True)

        top = ttk.Frame(root, padding=10); top.pack(fill="x")
        ttk.Label(top, text="Archivo Excel destino:").pack(side="left")
        from pathlib import Path as _Path
        self.path_var = tk.StringVar(value=str(_Path("scans_qr_tabulado.xlsx").resolve()))
        self.e_path = ttk.Entry(top, textvariable=self.path_var, width=70); self.e_path.pack(side="left", padx=6)
        ttk.Button(top, text="Cambiar...", command=self.choose_path).pack(side="left")

        self.csv_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(top, text="También guardar CSV", variable=self.csv_var).pack(side="left", padx=10)

        frm = ttk.Frame(root, padding=10); frm.pack(fill="both", expand=True)
        ttk.Label(frm, text="Coloca el cursor en la caja y escanea (el lector debe enviar ENTER al final).").pack(anchor="w")
        self.entry = ttk.Entry(frm, font=("Consolas", 13)); self.entry.pack(fill="x", pady=6); self.entry.focus_set()
        self.entry.bind("<Return>", self.on_enter)

        self.prev = tk.Text(frm, height=14, font=("Consolas", 10)); self.prev.pack(fill="both", expand=True)
        self.status = ttk.Label(root, text="Listo.", anchor="w"); self.status.pack(fill="x", padx=10, pady=(0,8))

        self.xlsx_path = Path(self.path_var.get())
        self.wb, self.ws = ensure_workbook(self.xlsx_path)
        self.csv_path = Path(str(self.xlsx_path.with_suffix(".csv")))
        self.ensure_csv_headers()

    def ensure_csv_headers(self):
        if not self.csv_var.get(): return
        p = self.csv_path
        if not p.exists():
            with open(p, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f); writer.writerow(COLUMNS)

    def choose_path(self):
        path = filedialog.asksaveasfilename(title="Guardar Excel", defaultextension=".xlsx", filetypes=[("Excel Workbook", "*.xlsx")])
        if not path: return
        self.path_var.set(path)
        from pathlib import Path as _Path
        self.xlsx_path = _Path(path)
        self.wb, self.ws = ensure_workbook(self.xlsx_path)
        self.csv_path = self.xlsx_path.with_suffix(".csv")
        self.ensure_csv_headers()
        self.status.config(text=f"Destino: {self.xlsx_path}")

    def on_enter(self, event=None):
        raw = self.entry.get(); self.entry.delete(0, "end")
        data = parse_scan(raw)

        import datetime as _dt
        now = _dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        row = {k:"" for k in COLUMNS}; row["timestamp"] = now; row["raw"] = raw

        for key in ("nombre_generico","concentracion","forma_farmaceutica","presentacion","control",
                    "registro_sanitario","lote","fecha_fabricacion","fecha_vencimiento",
                    "fabricante","importador","id_hospital","fecha_reempaque",
                    "conservacion","advertencias","normativa","url_qr"):
            if key in data: row[key] = str(data[key]).strip()

        self.ws.append([row[c] for c in COLUMNS]); self.wb.save(self.xlsx_path)
        if self.csv_var.get():
            with open(self.csv_path, "a", newline="", encoding="utf-8") as f:
                writer = csv.writer(f); writer.writerow([row[c] for c in COLUMNS])

        self.prev.delete("1.0","end")
        import json as _json
        pretty = _json.dumps(row, ensure_ascii=False, indent=2)
        self.prev.insert("end", f"✅ Escaneado y guardado en:\n  {self.xlsx_path}\n\n{pretty}\n")
        self.status.config(text=f"Guardado ({self.xlsx_path.name})  |  CSV: {self.csv_path.name if self.csv_var.get() else 'no'}")

        try:
            import ctypes; ctypes.windll.user32.MessageBeep(0x00000040)
        except Exception:
            pass

if __name__ == "__main__":
    try:
        import ctypes; ctypes.windll.user32.SetProcessDPIAware()
    except Exception:
        pass
    root = tk.Tk()
    style = ttk.Style()
    try: style.theme_use("clam")
    except Exception: pass
    App(root)
    root.mainloop()
